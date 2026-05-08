"""Write enrichment results to CSV, XLSX, JSONL, and summary JSON.

v0.4.1 — Clean output table: human-readable column names, logical grouping,
no internal/debug fields.  Raw data still available in JSONL.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

import pandas as pd

from email_osint_enricher.schemas import EnrichmentResult, RunSummary, OutputConfig

logger = logging.getLogger("enricher")


# ── Column mapping: internal field → clean header ───────────────────────────
# Ordered by groups.  Only these fields appear in CSV/XLSX.
# JSONL keeps full raw data for programmatic use.

COLUMN_MAP: list[tuple[str, str]] = [
    # ─── Identification ──────────────────────────────────────────────
    ("email",                               "Email"),
    ("applicantName",                       "Name"),
    ("applicantCountry",                    "Country"),
    ("applicantId",                         "Applicant ID"),
    ("externalId",                          "External ID"),

    # ─── Scores (the most important columns) ─────────────────────────
    ("final_enrichment_score",              "Final Score"),
    ("outreach_enrichment_tier",            "Tier"),
    ("manual_review_needed",                "Review Needed"),
    ("email_footprint_score",               "Footprint"),
    ("identity_confidence_score",           "Identity"),
    ("social_presence_score",               "Social Presence"),
    ("email_reputation_score",              "Reputation"),
    ("deliverability_score",                "Deliverability"),

    # ─── Holehe: registered services ─────────────────────────────────
    ("holehe_registered_services_count",    "Holehe: Services"),
    ("holehe_registered_services_list",     "Holehe: Service List"),

    # ─── Blackbird: profiles ─────────────────────────────────────────
    ("blackbird_email_profiles_count",      "Blackbird: Email Profiles"),
    ("blackbird_username_profiles_count",   "Blackbird: Username Profiles"),
    ("blackbird_profiles_list",             "Blackbird: Profile List"),

    # ─── Maigret: profiles ───────────────────────────────────────────
    ("maigret_profiles_count",              "Maigret: Profiles"),
    ("maigret_profiles_list",              "Maigret: Profile List"),

    # ─── Sherlock: profiles ──────────────────────────────────────────
    ("sherlock_profiles_count",             "Sherlock: Profiles"),
    ("sherlock_profiles_list",              "Sherlock: Profile List"),

    # ─── HudsonRock: cybercrime ──────────────────────────────────────
    ("hudsonrock_is_compromised",           "HudsonRock: Compromised"),
    ("hudsonrock_stealers_count",           "HudsonRock: Stealers"),
    ("hudsonrock_latest_compromise_date",   "HudsonRock: Last Compromise"),

    # ─── Gravatar: public profile ────────────────────────────────────
    ("gravatar_has_profile",                "Gravatar: Has Profile"),
    ("gravatar_display_name",               "Gravatar: Display Name"),
    ("gravatar_location",                   "Gravatar: Location"),
    ("gravatar_linked_accounts_count",      "Gravatar: Linked Accounts"),
    ("gravatar_linked_accounts",            "Gravatar: Account List"),
    ("gravatar_avatar_url",                 "Gravatar: Avatar URL"),

    # ─── Socialscan: platform checks ─────────────────────────────────
    ("socialscan_registered_count",         "Socialscan: Registered"),
    ("socialscan_registered_platforms",     "Socialscan: Platforms"),

    # ─── EmailRep: reputation ────────────────────────────────────────
    ("emailrep_reputation",                 "EmailRep: Reputation"),
    ("emailrep_suspicious",                 "EmailRep: Suspicious"),
    ("emailrep_references",                 "EmailRep: References"),

    # ─── Mosint (optional) ───────────────────────────────────────────
    ("mosint_social_signal",                "Mosint: Social"),
    ("mosint_breach_signal",                "Mosint: Breach"),
    ("mosint_domain_signal",                "Mosint: Domain"),

    # ─── EmailCrawlr (optional) ──────────────────────────────────────
    ("emailcrawlr_social_accounts_count",   "EmailCrawlr: Accounts"),
    ("emailcrawlr_deliverability",          "EmailCrawlr: Deliverability"),

    # ─── Phone extraction ────────────────────────────────────────────
    ("phone_candidates_count",              "Phones Found"),
    ("phone_candidate_best",                "Best Phone"),
    ("phone_candidates_list",               "Phone List"),

    # ─── Meta ────────────────────────────────────────────────────────
    ("total_profiles_found",                "Total Profiles"),
    ("enrichment_notes",                    "Notes"),
    ("recommended_next_action",             "Recommended Action"),
    ("source_provider",                     "Providers Used"),
    ("email_domain",                        "Domain"),
    ("email_type",                          "Domain Type"),
    ("processed_at",                        "Processed At"),
    ("status",                              "Status"),
]

# Internal fields to always strip from output dict
_STRIP_FIELDS = {
    "email_normalized", "input_row_id", "username_candidates",
    "error_message", "claim_value", "lead_score", "tier",
    # _checked / _success / _error / _raw_json_path / _confidence_score per provider
    # (we keep only the useful data columns above)
    "holehe_checked", "holehe_success", "holehe_social_services_count",
    "holehe_professional_services_count", "holehe_other_services_count",
    "holehe_raw_json_path", "holehe_confidence_score", "holehe_error",
    "blackbird_checked", "blackbird_success",
    "blackbird_report_path", "blackbird_raw_json_path",
    "blackbird_confidence_score", "blackbird_error",
    "maigret_checked", "maigret_success", "maigret_username_candidates",
    "maigret_report_path", "maigret_raw_json_path",
    "maigret_confidence_score", "maigret_error",
    "sherlock_checked", "sherlock_success",
    "sherlock_raw_json_path", "sherlock_confidence_score", "sherlock_error",
    "phone_extractor_checked", "phone_candidates_found",
    "phone_candidate_source_url", "phone_candidate_source_provider",
    "phone_candidate_context", "phone_candidate_confidence_score",
    "phone_extraction_error",
    "emailrep_checked", "emailrep_success",
    "emailrep_details_summary", "emailrep_risk_score",
    "emailrep_raw_json_path", "emailrep_error",
    "mosint_checked", "mosint_success", "mosint_services_used",
    "mosint_findings_count", "mosint_raw_json_path",
    "mosint_confidence_score", "mosint_error",
    "emailcrawlr_checked", "emailcrawlr_success",
    "emailcrawlr_social_accounts_list", "emailcrawlr_domain_emails_count",
    "emailcrawlr_raw_json_path", "emailcrawlr_confidence_score",
    "emailcrawlr_error",
    "hudsonrock_checked", "hudsonrock_success",
    "hudsonrock_total_corporate_services", "hudsonrock_total_user_services",
    "hudsonrock_compromised_dates", "hudsonrock_operating_systems",
    "hudsonrock_confidence_score", "hudsonrock_raw_json_path",
    "hudsonrock_error",
    "gravatar_checked", "gravatar_success",
    "gravatar_full_name", "gravatar_profile_url", "gravatar_about_me",
    "gravatar_confidence_score", "gravatar_raw_json_path", "gravatar_error",
    "socialscan_checked", "socialscan_success",
    "socialscan_not_registered_count",
    "socialscan_confidence_score", "socialscan_raw_json_path",
    "socialscan_error",
    # Scoring internals
    "provider_consensus_score", "conflict_risk_score",
    "merged_profiles_count",
}


def _result_to_clean_dict(r: EnrichmentResult) -> dict[str, Any]:
    """Convert EnrichmentResult to a clean dict with human-readable keys."""
    raw = r.model_dump()
    out: dict[str, Any] = {}
    field_to_header = {field: header for field, header in COLUMN_MAP}

    for field, header in COLUMN_MAP:
        val = raw.get(field)
        if val is None:
            val = ""
        # Clean booleans for readability in spreadsheets
        if isinstance(val, bool):
            val = "Yes" if val else "No"
        # Strip empty lists rendered as strings
        if val == "":
            pass  # keep empty
        out[header] = val

    return out


def _result_to_raw_dict(r: EnrichmentResult) -> dict:
    """Full dict for JSONL — keeps everything."""
    return r.model_dump()


def _auto_fit_xlsx(xlsx_path: Path, df: pd.DataFrame) -> None:
    """Auto-fit column widths and add header formatting for XLSX."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Header style
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            bottom=Side(style="thin", color="999999"),
        )

        # Group colors for visual separation
        group_colors = {
            "Identification": "E8F0FE",
            "Scores": "E6F4EA",
            "Holehe": "FFF3E0",
            "Blackbird": "FFF3E0",
            "Maigret": "FFF3E0",
            "Sherlock": "FFF3E0",
            "HudsonRock": "FCE4EC",
            "Gravatar": "F3E5F5",
            "Socialscan": "FFF3E0",
            "EmailRep": "E0F2F1",
            "Mosint": "E0F2F1",
            "EmailCrawlr": "E0F2F1",
            "Phone": "FFF8E1",
            "Meta": "F5F5F5",
        }

        def _get_group(header: str) -> str:
            if header in ("Email", "Name", "Country", "Applicant ID", "External ID"):
                return "Identification"
            if header in ("Final Score", "Tier", "Review Needed", "Footprint",
                          "Identity", "Social Presence", "Reputation", "Deliverability"):
                return "Scores"
            if header.startswith("Holehe"):
                return "Holehe"
            if header.startswith("Blackbird"):
                return "Blackbird"
            if header.startswith("Maigret"):
                return "Maigret"
            if header.startswith("Sherlock"):
                return "Sherlock"
            if header.startswith("HudsonRock"):
                return "HudsonRock"
            if header.startswith("Gravatar"):
                return "Gravatar"
            if header.startswith("Socialscan"):
                return "Socialscan"
            if header.startswith("EmailRep"):
                return "EmailRep"
            if header.startswith("Mosint"):
                return "Mosint"
            if header.startswith("EmailCrawlr"):
                return "EmailCrawlr"
            if "Phone" in header:
                return "Phone"
            return "Meta"

        for col_idx, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_idx)

            # Header cell
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Column width (auto-fit with min/max)
            max_len = len(str(col_name))
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 100),
                                    min_col=col_idx, max_col=col_idx):
                for c in row:
                    if c.value:
                        max_len = max(max_len, min(len(str(c.value)), 60))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)

            # Group color for data rows
            group = _get_group(col_name)
            if group in group_colors:
                data_fill = PatternFill(
                    start_color=group_colors[group],
                    end_color=group_colors[group],
                    fill_type="solid",
                )
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                        min_col=col_idx, max_col=col_idx):
                    for c in row:
                        c.fill = data_fill
                        c.border = thin_border
                        c.alignment = Alignment(vertical="center", wrap_text=False)

        # Freeze top row + first column (Email)
        ws.freeze_panes = "B2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        wb.save(xlsx_path)
        logger.info(f"XLSX formatted: {xlsx_path}")

    except Exception as e:
        logger.warning(f"XLSX formatting skipped: {e}")


def write_results(
    results: list[EnrichmentResult],
    errors: list[EnrichmentResult],
    summary: RunSummary,
    output_dir: str | Path,
    config: OutputConfig,
) -> dict[str, str]:
    """Write all output files and return paths."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    paths: dict[str, str] = {}

    if results:
        # Clean records for CSV/XLSX
        clean_records = [_result_to_clean_dict(r) for r in results]
        df = pd.DataFrame(clean_records)

        # Drop columns where ALL values are empty / "No" / 0
        # (provider wasn't run → don't show its columns)
        cols_to_drop = []
        for col in df.columns:
            vals = df[col]
            if vals.apply(lambda x: x in ("", "No", 0, "0", None)).all():
                cols_to_drop.append(col)
        # Never drop identification or score columns
        keep_always = {
            "Email", "Name", "Country", "Final Score", "Tier",
            "Review Needed", "Status",
        }
        cols_to_drop = [c for c in cols_to_drop if c not in keep_always]
        df = df.drop(columns=cols_to_drop)

        if config.write_csv:
            csv_path = output_dir / "enriched_results.csv"
            df.to_csv(csv_path, index=False, encoding="utf-8-sig")
            paths["csv"] = str(csv_path)
            logger.info(f"Written: {csv_path}")

        if config.write_xlsx:
            xlsx_path = output_dir / "enriched_results.xlsx"
            df.to_excel(xlsx_path, index=False, engine="openpyxl")
            _auto_fit_xlsx(xlsx_path, df)
            paths["xlsx"] = str(xlsx_path)
            logger.info(f"Written: {xlsx_path}")

        if config.write_jsonl:
            jsonl_path = output_dir / "enriched_results.jsonl"
            with open(jsonl_path, "w", encoding="utf-8") as f:
                for r in results:
                    f.write(r.model_dump_json() + "\n")
            paths["jsonl"] = str(jsonl_path)
            logger.info(f"Written: {jsonl_path}")
    else:
        logger.warning("No results to write")

    # Errors CSV
    if errors:
        err_records = [{"Email": r.email, "Error": r.error_message, "Status": r.status.value}
                       for r in errors]
        err_df = pd.DataFrame(err_records)
        err_path = output_dir / "errors.csv"
        err_df.to_csv(err_path, index=False, encoding="utf-8-sig")
        paths["errors"] = str(err_path)
        logger.info(f"Written: {err_path}")

    # Run summary
    summary_path = output_dir / "run_summary.json"
    summary_path.write_text(
        summary.model_dump_json(indent=2),
        encoding="utf-8",
    )
    paths["summary"] = str(summary_path)
    logger.info(f"Written: {summary_path}")

    return paths
